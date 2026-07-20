from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import re
import zipfile
import json
import uuid
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Optional

import bcrypt
import jwt
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from sqlalchemy import create_engine, Column, String, Integer, Text, DateTime, ForeignKey
from sqlalchemy.orm import declarative_base, sessionmaker, relationship, Session

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Database (SQL via SQLAlchemy)
# ---------------------------------------------------------------------------
DATABASE_URL = os.environ["DATABASE_URL"]
engine = create_engine(
    DATABASE_URL,
    connect_args={"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {},
)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)
Base = declarative_base()


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, autoincrement=True)
    username = Column(String(80), unique=True, nullable=False, index=True)
    password_hash = Column(String(255), nullable=False)
    name = Column(String(120), nullable=False, default="")
    role = Column(String(30), nullable=False, default="user")
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    jobs = relationship("ConversionJob", back_populates="user", cascade="all, delete-orphan")


class ConversionJob(Base):
    __tablename__ = "conversion_jobs"
    id = Column(String(36), primary_key=True, default=lambda: str(uuid.uuid4()))
    user_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    title = Column(String(200), nullable=False, default="Conversion")
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    file_count = Column(Integer, default=0)
    row_count = Column(Integer, default=0)
    filenames = Column(Text, default="[]")       # JSON list of processed filenames
    errors = Column(Text, default="[]")           # JSON list of {filename, error}
    rows = Column(Text, default="[]")             # JSON list of extracted row dicts
    user = relationship("User", back_populates="jobs")


Base.metadata.create_all(bind=engine)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
JWT_ALGORITHM = "HS256"


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def create_access_token(user_id: int, username: str) -> str:
    payload = {
        "sub": str(user_id),
        "username": username,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def get_current_user(request: Request, db: Session = Depends(get_db)) -> User:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        user = db.query(User).filter(User.id == int(payload["sub"])).first()
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


# ---------------------------------------------------------------------------
# Word parsing
# ---------------------------------------------------------------------------
HEADING_COLUMNS = ["Title", "Heading 1", "Heading 2"]


def _style_name(paragraph) -> str:
    try:
        return (paragraph.style.name or "").strip()
    except Exception:
        return ""


def _num_level(paragraph):
    """Return the list numbering indent level (0 = main, 1 = sub) or None."""
    pPr = paragraph._p.pPr
    if pPr is None or pPr.numPr is None:
        return None
    ilvl = pPr.numPr.ilvl
    return int(ilvl.val) if ilvl is not None else 0


def extract_custom_props(file_bytes: bytes) -> dict:
    """Read SharePoint documentManagement metadata embedded in the .docx
    (customXml/itemN.xml). The item index varies between files, so scan all."""
    props = {"Area": "", "UserCode": "", "FileNo": "", "Title": ""}
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            for name in z.namelist():
                if name.startswith("customXml/item") and name.endswith(".xml"):
                    content = z.read(name).decode("utf-8", errors="ignore")
                    for tag in ("FileNo", "Area", "UserCode"):
                        if not props[tag]:
                            m = re.search(r"<%s\b[^>]*>([^<]*)</%s>" % (tag, tag), content)
                            if m:
                                props[tag] = m.group(1).strip()
                    if not props["Title"]:
                        m = re.search(r"<Word_x002d_Title\b[^>]*>([^<]*)</Word_x002d_Title>", content)
                        if m:
                            props["Title"] = m.group(1).strip()
    except Exception:
        pass
    return props


def _split_at(text: str):
    """Split a fact on '@'. Returns (body, keyword, area).
    keyword/area are None when the corresponding '@' segment is absent."""
    parts = text.split("@")
    body = parts[0]
    keyword = parts[1].strip() if len(parts) >= 2 else None
    area = "@".join(parts[2:]).strip() if len(parts) >= 3 else None
    return body, keyword, area


FACT_COLUMNS = [
    "Fact#", "Facts", "Title", "Scope", "File Name",
    "Keyword", "Area", "Team Area", "Reviewer",
]


def parse_docx(file_bytes: bytes, filename: str) -> List[dict]:
    """Extract observation Facts from a WANO Field Note .docx into structured rows.

    - Title  : SharePoint 'Word-Title' custom prop (fallback: first Heading 1).
    - Scope  : first non-empty paragraph after the 'SCOPE' heading.
    - Facts  : List Paragraph items under 'OBSERVATIONS'. Level 0 => main fact
               (1, 2, 3...), level 1 => sub-fact (1.a, 1.b...). Numbering resets
               per file. The text after the first '@' (Keyword) and second '@'
               (Area) is stripped from the body and a
               '#{Reviewer}-{FileNo}_{Fact#}' tag is appended.
    - Team Area = custom 'Area'; Reviewer = Area + UserCode.
    """
    document = Document(io.BytesIO(file_bytes))
    cp = extract_custom_props(file_bytes)
    area = cp["Area"]
    user_code = cp["UserCode"]
    file_no = cp["FileNo"]
    reviewer = f"{area}{user_code}"

    paras = document.paragraphs

    title = cp["Title"]
    if not title:
        for p in paras:
            if _style_name(p) == "Heading 1" and p.text.strip():
                title = p.text.strip()
                break

    scope = ""
    for i, p in enumerate(paras):
        if _style_name(p).startswith("Heading") and p.text.strip().upper() == "SCOPE":
            for q in paras[i + 1:]:
                if q.text.strip():
                    scope = q.text.strip()
                    break
            break

    rows: List[dict] = []
    main = 0
    sub_ord = 0
    cur_kw = ""
    cur_area = ""
    in_obs = False

    for p in paras:
        style = _style_name(p)
        text = (p.text or "").strip()
        if style.startswith("Heading") and text.upper().startswith("OBSERVATION"):
            in_obs = True
            continue
        if not in_obs or style != "List Paragraph" or not text:
            continue

        lvl = _num_level(p)
        if lvl is None:
            lvl = 0
        body, kw, ar = _split_at(text)

        if lvl == 0:
            main += 1
            sub_ord = 0
            fact_no = str(main)
            cur_kw = kw or ""
            cur_area = ar or ""
            row_kw, row_area = cur_kw, cur_area
        else:
            fact_no = f"{main}.{chr(ord('a') + sub_ord)}"
            sub_ord += 1
            row_kw = kw if kw is not None else cur_kw
            row_area = ar if ar is not None else cur_area

        body_clean = body.rstrip()
        suffix = f"#{reviewer}-{file_no}_{fact_no}"
        facts_text = f"{body_clean} {suffix}" if body_clean else suffix

        rows.append({
            "Fact#": fact_no,
            "Facts": facts_text,
            "Title": title,
            "Scope": scope,
            "File Name": filename,
            "Keyword": row_kw,
            "Area": row_area,
            "Team Area": area,
            "Reviewer": reviewer,
        })

    return rows


EXCEL_COLUMNS = FACT_COLUMNS
EXCEL_HEADERS = {k: k for k in FACT_COLUMNS}


def build_excel(rows: List[dict], filenames: List[str]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facts"

    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    header_fill = PatternFill("solid", fgColor="002FA7")
    for idx, key in enumerate(FACT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, start=2):
        for c, key in enumerate(FACT_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(key, ""))
            if key in ("Facts", "Scope", "Title"):
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "Fact#": 8, "Facts": 80, "Title": 38, "Scope": 38, "File Name": 22,
        "Keyword": 20, "Area": 16, "Team Area": 12, "Reviewer": 12,
    }
    for idx, key in enumerate(FACT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(key, 16)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------
app = FastAPI()
api_router = APIRouter(prefix="/api")


class LoginRequest(BaseModel):
    username: str
    password: str


class UserOut(BaseModel):
    id: int
    username: str
    name: str
    role: str


@api_router.get("/")
async def root():
    return {"message": "Word to Excel API"}


@api_router.post("/auth/login")
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    username = payload.username.strip().lower()
    user = db.query(User).filter(User.username == username).first()
    if not user or not verify_password(payload.password, user.password_hash):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    token = create_access_token(user.id, user.username)
    return {
        "token": token,
        "user": {"id": user.id, "username": user.username, "name": user.name, "role": user.role},
    }


@api_router.get("/auth/me", response_model=UserOut)
def me(current: User = Depends(get_current_user)):
    return UserOut(id=current.id, username=current.username, name=current.name, role=current.role)


@api_router.post("/convert")
async def convert(
    files: List[UploadFile] = File(...),
    current: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if len(files) > 200:
        raise HTTPException(status_code=400, detail="Maximum 200 files allowed per conversion")
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    all_rows: List[dict] = []
    processed: List[str] = []
    errors: List[dict] = []

    for f in files:
        fname = f.filename or "unknown.docx"
        try:
            content = await f.read()
            if not fname.lower().endswith(".docx"):
                errors.append({"filename": fname, "error": "Unsupported format. Only .docx is supported (.doc must be converted first)."})
                continue
            rows = parse_docx(content, fname)
            all_rows.extend(rows)
            processed.append(fname)
        except Exception as e:
            logger.exception("Failed to parse %s", fname)
            errors.append({"filename": fname, "error": str(e)})

    if not processed:
        raise HTTPException(status_code=400, detail="No valid .docx files could be processed. " + (errors[0]["error"] if errors else ""))

    job = ConversionJob(
        user_id=current.id,
        title=f"{len(processed)} file(s) - {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')}",
        file_count=len(processed),
        row_count=len(all_rows),
        filenames=json.dumps(processed),
        errors=json.dumps(errors),
        rows=json.dumps(all_rows),
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    return {
        "job_id": job.id,
        "file_count": job.file_count,
        "row_count": job.row_count,
        "filenames": processed,
        "errors": errors,
        "columns": [EXCEL_HEADERS.get(k, k) for k in EXCEL_COLUMNS],
        "preview": all_rows[:100],
    }


@api_router.get("/jobs")
def list_jobs(current: User = Depends(get_current_user), db: Session = Depends(get_db)):
    jobs = db.query(ConversionJob).filter(ConversionJob.user_id == current.id).order_by(ConversionJob.created_at.desc()).all()
    return [
        {
            "job_id": j.id,
            "title": j.title,
            "created_at": j.created_at.isoformat() if j.created_at else "",
            "file_count": j.file_count,
            "row_count": j.row_count,
            "filenames": json.loads(j.filenames or "[]"),
            "errors": json.loads(j.errors or "[]"),
        }
        for j in jobs
    ]


@api_router.get("/jobs/{job_id}")
def get_job(job_id: str, current: User = Depends(get_current_user), db: Session = Depends(get_db)):
    job = db.query(ConversionJob).filter(ConversionJob.id == job_id, ConversionJob.user_id == current.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    rows = json.loads(job.rows or "[]")
    return {
        "job_id": job.id,
        "title": job.title,
        "file_count": job.file_count,
        "row_count": job.row_count,
        "filenames": json.loads(job.filenames or "[]"),
        "errors": json.loads(job.errors or "[]"),
        "columns": [EXCEL_HEADERS.get(k, k) for k in EXCEL_COLUMNS],
        "preview": rows[:100],
    }


@api_router.get("/jobs/{job_id}/download")
def download_job(job_id: str, current: User = Depends(get_current_user), db: Session = Depends(get_db)):
    job = db.query(ConversionJob).filter(ConversionJob.id == job_id, ConversionJob.user_id == current.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    rows = json.loads(job.rows or "[]")
    filenames = json.loads(job.filenames or "[]")
    data = build_excel(rows, filenames)
    fname = f"word-to-excel-{job.id[:8]}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api_router.delete("/jobs/{job_id}")
def delete_job(job_id: str, current: User = Depends(get_current_user), db: Session = Depends(get_db)):
    job = db.query(ConversionJob).filter(ConversionJob.id == job_id, ConversionJob.user_id == current.id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    db.delete(job)
    db.commit()
    return {"ok": True}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


def seed_admin():
    db = SessionLocal()
    try:
        username = os.environ.get("ADMIN_USERNAME", "admin").strip().lower()
        password = os.environ.get("ADMIN_PASSWORD", "admin123")
        existing = db.query(User).filter(User.username == username).first()
        if existing is None:
            db.add(User(username=username, password_hash=hash_password(password), name="Administrator", role="admin"))
            db.commit()
            logger.info("Seeded admin user '%s'", username)
        elif not verify_password(password, existing.password_hash):
            existing.password_hash = hash_password(password)
            db.commit()
    finally:
        db.close()


@app.on_event("startup")
def on_startup():
    seed_admin()
